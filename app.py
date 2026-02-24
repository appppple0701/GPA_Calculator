#模組和模板
#主要邏輯與前端
import gpa
import streamlit as st
import pandas as pd

#模板下載
from io import BytesIO     
from openpyxl.comments import Comment

#這個shit會回傳一個模板
def build_template_xlsx() -> bytes:
#courses sheet
    df_courses = pd.DataFrame(columns=[
        "term",
        "course",
        "score",
        "credit",
        "count_gpa",
    ])

#ranks sheet
    df_ranks = pd.DataFrame(columns=[
        "term",
        "class_rank",
        "class_size",
        "dept_rank",
        "dept_size",
        "sem_grade",
    ])

    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df_courses.to_excel(writer, sheet_name="courses", index=False)
        df_ranks.to_excel(writer, sheet_name="ranks", index=False)

        wb = writer.book

        #courses的註解
        ws = wb["courses"]
        ws["A1"].comment = Comment(
            "學年度-學期，例如 2024-1；1=上學期(Fall)，2=下學期(Spring)",
            "GPA Calculator"
        )
        ws["B1"].comment = Comment("課程名稱（文字）", "GPA Calculator")
        ws["C1"].comment = Comment("分數（數字，滿分 100.0）", "GPA Calculator")
        ws["D1"].comment = Comment("學分數（數字）", "GPA Calculator")
        ws["E1"].comment = Comment(
            "是否列入 GPA 計算（選填；1=列入；2=不列入）",
            "GPA Calculator"
        )
        ws.freeze_panes = "A2"

        #ranks的註解
        ws2 = wb["ranks"]
        ws2["A1"].comment = Comment(
            "學年度-學期，需與 courses 的 term 對齊",
            "GPA Calculator"
        )
        ws2["B1"].comment = Comment("班排名（數字，1 表示第一名）", "GPA Calculator")
        ws2["C1"].comment = Comment("班級人數（數字）", "GPA Calculator")
        ws2["D1"].comment = Comment("系排名（選填）", "GPA Calculator")
        ws2["E1"].comment = Comment("系人數（選填）", "GPA Calculator")
        ws2["F1"].comment = Comment("學期成績（數字）", "GPA Calculator")
        ws2.freeze_panes = "A2"

    bio.seek(0)
    return bio.getvalue()

#讓高師大學生也能拿到excel檔
def build_result_xlsx(df_courses: pd.DataFrame, df_ranks: pd.DataFrame) -> bytes:
    bio = BytesIO()

    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df_courses.to_excel(writer, sheet_name="courses", index=False)
        df_ranks.to_excel(writer, sheet_name="ranks", index=False)

    bio.seek(0)
    return bio.getvalue()
#高師大學生快速貼上
import re
from io import StringIO
from datetime import datetime

#把原始資料轉成兩個sheets
def parse_nknu_paste_text(raw: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    raw = (raw or "").strip()
    if not raw:
        raise ValueError("Empty paste text")

    term_re = re.compile(r"(\d{3})\s*學年度.*第\s*([12])\s*學期")
    courses = []
    ranks = []

    current_term = None
    sem_grade = None
    class_rank = None
    class_size = None

    def flush_rank():
        nonlocal sem_grade, class_rank, class_size, current_term
        if current_term is None:
            return
        ranks.append({
            "term": current_term,
            "class_rank": class_rank,
            "class_size": class_size,
            "dept_rank": None,
            "dept_size": None,
            "sem_grade": sem_grade,
        })

    def split_cols(line: str) -> list[str]:
        # 先用 tab 拆；若沒有 tab，再用 2 個以上空白拆
        if "\t" in line:
            cols = [c.strip() for c in line.split("\t")]
        else:
            cols = [c.strip() for c in re.split(r"\s{2,}", line.strip())]
        return [c for c in cols if c != ""]

    lines = raw.splitlines()
    for line in lines:
        line = line.strip()
        if not line:
            continue

        # --- 學期標題 ---
        m = term_re.search(line)
        if m:
            flush_rank()
            roc = int(m.group(1))
            sem = int(m.group(2))
            year = roc + 1911
            current_term = f"{year}-{sem}"
            sem_grade = None
            class_rank = None
            class_size = None
            continue

        if current_term is None:
            continue

        #摘要
        if "學期平均" in line:
            cols = split_cols(line)
            # 常見：學期平均 ... 83.21/85
            last = cols[-1] if cols else ""
            last = last.replace("／", "/").replace(" ", "")
            try:
                sem_grade = float(last.split("/")[0])
            except:
                sem_grade = None
            continue

        #學期名次
        if "學期名次" in line:
            cols = split_cols(line)
            last = cols[-1] if cols else ""
            last = last.replace("／", "/").replace(" ", "")
            parts = last.split("/")
            if len(parts) >= 2:
                try:
                    class_rank = int(float(parts[0]))
                except:
                    class_rank = None
                try:
                    class_size = int(float(parts[1]))
                except:
                    class_size = None
            continue

        #表頭略過
        if line.startswith("科目名稱"):
            continue

        #課程列
        cols = split_cols(line)
        # 典型欄位：科目名稱 | 學分 | 歸類 | 必選修 | 分數
        # 我們只需要：course=0, credit=1, score=最後一欄
        if len(cols) < 2:
            continue

        course = cols[0]
        credit_raw = cols[1]
        score_raw = cols[-1]

        # 避免把摘要列當課程
        if any(k in course for k in ["修習學分", "學期平均", "學期名次"]):
            continue

        # credit: 可能是 -3（不列入本系學分），我們先轉數字再 abs
        credit_val = None
        try:
            credit_val = abs(float(str(credit_raw).replace("－", "-")))
        except:
            credit_val = None

        # score: 可能是 未送/空白
        score_val = None
        s = str(score_raw).strip()
        if s in ("未送", "-", ""):
            score_val = None
        else:
            try:
                score_val = float(s)
            except:
                score_val = None

        courses.append({
            "term": current_term,
            "course": course,
            "score": score_val,
            "credit": credit_val,
            "count_gpa": 1,   # 先預設都列入，使用者可在前端取消勾選
        })

    flush_rank()

    df_courses = pd.DataFrame(courses, columns=["term", "course", "score", "credit", "count_gpa"])
    df_ranks = pd.DataFrame(ranks, columns=["term", "class_rank", "class_size", "dept_rank", "dept_size", "sem_grade"])
    return df_courses, df_ranks

#前端'''
#------------------------網頁標題-------------------------------------
st.title("GPA CALCULATOR")

#選擇資料匯入方式
#用側邊欄選擇
load = st.sidebar.radio("請選擇資料上傳方式", ["上傳Excel", "高師大學生快速匯入"])
#選擇模板匯入
if load == "上傳Excel":
    st.subheader("檔案上傳")
    file = st.file_uploader(label= "請上傳資料或下載模板")
    if file is not None:     #已上傳檔案 則顯示分析
        df_courses, df_ranks = gpa.load_grade_file_auto(file)
    else:     #未上傳檔案則顯示教學(用砂小expander放)
        #st.caption("檔案格式與上傳示範")
        #模板下載按鈕
        st.download_button(
            label = "下載GPA Excel模板",
            data = build_template_xlsx(),
            file_name = "成績單模板.xlsx",
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        #顯示教學
        with st.expander("模板匯入 (適用於所有學校)"):
            image1, image2 = st.columns(2)
            image3, image4 = st.columns(2)
            image1.image("image/下載模板.png",caption= "(1)點擊 \"下載GPA Excel模板\"")
            image2.image("image/填寫模板1.png", caption="(2)依照格式填寫sheet1 \"courses\" (一行一筆資料)")
            image3.image("image/填寫模板2.png", caption="(3)依照格式填寫sheet2 \"ranks\" (一行一筆資料)")
            image4.image("image/上傳資料.png", caption="(4)將填好的Excel下載 並且上傳至 \"GPA CALCULATOR\"")
        st.stop()

#選擇高師快速貼上
elif load == "高師大學生快速匯入":
    st.subheader("資料貼上區")
    #讀文字進去 存在raw變數
    raw = st.text_area("貼上高師大歷年成績（Ctrl+V）", height=220)

    if not raw.strip():     #如果掃上來是空的 顯示教學按鈕 
        with st.expander("高師大快速匯入 (免整理格式)"):
            st.image("image/歷年成績查詢.png",caption="(1)高師大學生請至\"單一登入平台\"的\"歷年成績查詢\"複製資料")
            st.image("image/複製資料.png",caption="(2)將複製好的資料完整貼上\"資料上傳區\"")
            #st.image("image/貼到excel.png",caption="(2)將複製好的資料完整貼上\"資料上傳區\"")
            st.stop()
    try:
        df_courses, df_ranks = parse_nknu_paste_text(raw)
    except Exception as e:
        st.error(f"解析失敗：{e}")
        st.stop()

#-------------------課程預覽區-------------------------------------
st.subheader("課程預覽與勾選")
st.caption("請勾選要列入計算之課程")
df_courses["include"] = df_courses["count_gpa"].fillna(1).astype(int).eq(1)
#根據count_gpa這個欄位，產生內部用的欄位include
#fillna(1) : 如果使用者漏填的話自動補1
#astype(int) : 如果使用者選錯儲存格格式，則自斷轉換成int
#eq(1) : 把儲存格中的int轉成布林值，1=True；0=False  

df_courses_edit = st.data_editor(
    df_courses,
    column_config = {"include" : st.column_config.CheckboxColumn("列入GPA計算")},
    disabled = ["term", "course", "score", "credit"]
)
#根據使用者勾選的結果，產生新的資料 df_course_edit 
#st.data_editor : 允許使用者在前端編輯，並且產生出新的df
#column_config : 把某欄位顯示成想要的樣子
#disabled : 絕對不要被編輯到的欄位

df_courses_calc = df_courses_edit[df_courses_edit["include"]].copy()
#產生一張新的df(複製出來的，不要影響原資料)


#--------------------結果分析區----------------------------------
st.subheader("分析結果")
system = st.radio("請選擇GPA制度",("4.0", "4.3"))

#產生歷年gpa的df
terms = sorted(df_courses_calc["term"].unique())
gpas = [
    gpa.calculate_gpa(
        df_courses_calc[df_courses_calc["term"] == t],
        system
    )
    for t in terms
    ]
df_gpa = pd.DataFrame({
    "term" : terms,
    "gpa" : gpas
})

#排版用---------------------------------------------------------------------------------------------------
left_column1, right_column1 = st.columns(2)
left_column2, mid_column2, right_column2 = st.columns(3)
left_column3, right_column3 = st.columns(2)
#排版用---------------------------------------------------------------------------------------------------

#----------------------GPA相關資料''''''''''''''''''''''''''''''''''
#顯示歷年GPA
left_column1.write("歷年GPA結果")
left_column1.write(df_gpa)
left_column1.subheader(f"平均GPA結果 : {gpa.calculate_gpa(df_courses_calc, system)}")
#left_column1.write(gpa.calculate_gpa(df_courses_calc, system))

#顯示GPA折線圖
right_column1.write("GPA折線圖")
#right_column1.write(gpa.calculate_gpa(df_courses_calc, system))
right_column1.line_chart(df_gpa, x = "term", y = "gpa")


#-----------------------排名相關資料'''''''''''''''''''''''
#產生歷年排名的df
terms = sorted(df_ranks["term"].unique())

class_prs = [
    gpa.calculate_pr(
        df_ranks[df_ranks["term"] == t]["class_rank"].iloc[0],
        df_ranks[df_ranks["term"] == t]["class_size"].iloc[0]
    )
    for t in terms
]

dept_prs = []
for t in terms:
    row = df_ranks[df_ranks["term"] == t].iloc[0]
    if pd.notna(row.get("dept_rank")) and pd.notna(row.get("dept_size")):
        dept_prs.append(gpa.calculate_pr(row["dept_rank"], row["dept_size"]))
    else:
        dept_prs.append(None)

df_rank = pd.DataFrame({
    "term": terms,
    "class_rank": [df_ranks[df_ranks["term"] == t]["class_rank"].iloc[0] for t in terms],
    "class_pr": class_prs,
    "dept_rank": [df_ranks[df_ranks["term"] == t].get("dept_rank", pd.Series([None])).iloc[0] for t in terms],
    "dept_pr": dept_prs,
})

#顯示歷年排名
left_column2.write("歷年排名結果")
left_column2.write(df_rank)

#顯示排名折線圖
mid_column2.write("排名折線圖(Pr)")
mid_column2.line_chart(df_rank, x = "term", y = ["class_pr", "dept_pr"])

#顯示排名折線圖
import altair as alt     

right_column2.write("排名折線圖（數字越小代表表現越好）")

chart = alt.Chart(df_rank).mark_line(point=True).encode(
    x=alt.X("term:N", title="學期"),
    y=alt.Y(
        "class_rank:Q",
        title="班排名（越小越好）",
        scale=alt.Scale(reverse=True)   
    ),
    tooltip=["term", "class_rank"]
)

right_column2.altair_chart(chart, use_container_width=True)

#-------------------------學期成績相關資料'''''''''''''''''''''''''''''''-
#建立歷年學期成績df
df_sem_grade = pd.DataFrame({
    "term" : df_ranks["term"],
    "sem_grade" : df_ranks["sem_grade"]
})

#顯示歷年學期成績
left_column3.write("歷年學期成績")
left_column3.write(df_sem_grade)

#顯示學期成績折線圖
right_column3.write("學期成績折線圖")
right_column3.line_chart(df_sem_grade, x = "term", y = "sem_grade")

st.subheader("成績單下載")

st.download_button(
    label="下載成績單 (Excel格式)",
    data=build_result_xlsx(df_courses_calc, df_ranks),
    file_name="GPA成績單.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

#st.sidebar.markdown("### 📝 使用回饋")
#st.sidebar.markdown("[👉 點我填寫回饋表單](https://forms.gle/2ZFEE3JVatDS5RYu9)")